import os

import openpyxl
from persiantools import characters, digits

from utils import project_variables


def get_path(file_name, directory):
    return os.path.join(directory, file_name + ".xlsx")


def fix_lms_teachers_name(directory):
    path = get_path(project_variables.TEACHERS_EXCEL_NAME, directory)
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if type(cell.value) == str:
                    teacher_name = cell.value
                    if 'سیده' in teacher_name:
                        teacher_names = teacher_name.split('سیده')
                        teacher_name = 'سیده ' + teacher_names[1].strip()
                    elif 'سید' in teacher_name:
                        teacher_names = teacher_name.split('سید')
                        teacher_name = 'سید ' + teacher_names[1].strip()
                    cell.value = teacher_name
    workbook.save(path)


def make_name_correct(name: str):
    if name.isspace() or name == '':
        return name
    if 'سيده' in name:
        parts = name.split('سيده')
        name = 'سيده ' + parts[1].strip() + ' ' + parts[0].strip()
    elif 'سيد' in name:
        parts = name.split('سيد')
        name = 'سيد ' + parts[1].strip() + ' ' + parts[0].strip()
    else:
        parts = name.split()
        if 'سادات' in name:
            name = parts[-2] + ' ' + parts[-1] + ' ' + ' '.join(parts[:-2])
        else:
            name = parts[-1] + ' ' + ' '.join(parts[:-1])
    return name


def clean_data(file_name, directory):
    replace_arabian_with_persian(file_name, directory)
    replace_slash_with_dot_in_numbers(file_name, directory)


def replace_arabian_with_persian(file_name, directory):
    print(os.getcwd())
    path = get_path(file_name, directory)
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if type(cell.value) == str:
                    cell.value = characters.ar_to_fa(cell.value)
                    if not cell.value.isalpha():
                        cell.value = digits.fa_to_en(cell.value)
    workbook.save(path)


def replace_slash_with_dot_in_numbers(file_name, directory):
    path = get_path(file_name, directory)
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if type(cell.value) == str and '0/5' in cell.value:
                    cell.value = cell.value.replace('/', '.')
    workbook.save(path)
